import pandas as pd
import re
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# -------------------------------------------
# Decomposição do Código de Horário UFSC
# -------------------------------------------
def quebrar_horario_ufsc(string_horario):
    mapa_dias = {
        '2': 'Segunda-feira', '3': 'Terça-feira', '4': 'Quarta-feira',
        '5': 'Quinta-feira', '6': 'Sexta-feira', '7': 'Sábado'
    }

    blocos_matutinos = ["07:30 - 08:20", "08:20 - 09:10", "09:10 - 10:10", "10:10 - 11:00", "11:00 - 11:50"]
    blocos_vespertinos = ["13:30 - 14:20", "14:20 - 15:10", "15:10 - 16:20", "16:20 - 17:10", "17:10 - 18:00"]
    blocos_noturnos = ["18:30 - 19:20", "19:20 - 20:20", "20:20 - 21:10", "21:10 - 22:00"]

    horarios_explodidos = []
    sub_horarios = str(string_horario).split(" | ")

    for sub in sub_horarios:
        match = re.match(r'(\d)\.(\d{4})-(\d)', sub.strip())
        if match:
            dia_num = match.group(1)
            inicio = match.group(2)
            creditos = int(match.group(3))

            dia_semana = mapa_dias.get(dia_num, "Desconhecido")
            hora_formatada_inicio = f"{inicio[:2]}:{inicio[2:]}"

            if inicio >= "1830":
                grade_blocos = blocos_noturnos
                mapa_indices = {"18:30": 0, "19:20": 1, "20:20": 2, "21:10": 3}
            elif inicio >= "1330":
                grade_blocos = blocos_vespertinos
                mapa_indices = {"13:30": 0, "14:20": 1, "15:10": 2, "16:20": 3, "17:10": 4}
            else:
                grade_blocos = blocos_matutinos
                mapa_indices = {"07:30": 0, "08:20": 1, "09:10": 2, "10:10": 3, "11:00": 4}

            idx_inicio = mapa_indices.get(hora_formatada_inicio, 0)

            for c in range(creditos):
                if idx_inicio + c < len(grade_blocos):
                    horarios_explodidos.append({
                        "Dia": dia_semana,
                        "Horário": grade_blocos[idx_inicio + c]
                    })

    return horarios_explodidos


# -------------------------------------------
# Geração da Matriz de Grade Horária Por Fase
# -------------------------------------------
def gerar_grade_horaria_fase(df_fase):
    todos_dias = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
    todos_horarios = [
        "07:30 - 08:20", "08:20 - 09:10", "09:10 - 10:10", "10:10 - 11:00", "11:00 - 11:50",
        "13:30 - 14:20", "14:20 - 15:10", "15:10 - 16:20", "16:20 - 17:10", "17:10 - 18:00",
        "18:30 - 19:20", "19:20 - 20:20", "20:20 - 21:10", "21:10 - 22:00"
    ]

    linhas_grade = []

    for _, linha_materia in df_fase.iterrows():
        string_horario = str(linha_materia["Horário"])

        if string_horario and string_horario not in ["A definir", "N/A"]:
            blocos_tempo = quebrar_horario_ufsc(string_horario)
            for bloco in blocos_tempo:
                linhas_grade.append({
                    "Dia": bloco["Dia"],
                    "Horário": bloco["Horário"],
                    "Identificador": f"{linha_materia['Código da Disciplina']} ({linha_materia['Turma']})"
                })

    if not linhas_grade:
        return pd.DataFrame("-", index=todos_horarios, columns=todos_dias)

    df_explodido = pd.DataFrame(linhas_grade)
    df_agrupado = df_explodido.groupby(["Horário", "Dia"])["Identificador"].apply(lambda x: " / ".join(x)).reset_index()

    grade_pivotada = df_agrupado.pivot(index="Horário", columns="Dia", values="Identificador")
    grade_final = grade_pivotada.reindex(index=todos_horarios, columns=todos_dias).fillna("-")

    return grade_final


# -------------------------------------------
# Exportação Formatada para o Excel (.xlsx)
# -------------------------------------------
def exportar_relatorios_finais(dados_finais: pd.DataFrame, nome_arquivo_excel: str):
    print(f"[Excel] Gerando arquivo Excel final com colunas divididas...")

    with pd.ExcelWriter(nome_arquivo_excel, engine="openpyxl") as writer:
        dados_finais.to_excel(writer, sheet_name="Cadastro Geral", index=False)

        for fase_num in range(1, 7):
            fase_str = f"{fase_num}ª Fase"
            
            # ✅ CORREÇÃO AQUI: Usa str.contains para incluir tanto '5ª Fase' quanto '5ª Fase (Optativa)'
            df_fase = dados_finais[dados_finais["Fase"].astype(str).str.contains(fase_str, na=False)]

            if not df_fase.empty:
                grade_semanal = gerar_grade_horaria_fase(df_fase)

                if not grade_semanal.empty:
                    grade_semanal.to_excel(writer, sheet_name=f"{fase_num}ª fase", index=True)

                    wb = writer.book
                    ws = wb[f"{fase_num}ª fase"]
                    ws.cell(row=1, column=1).value = "Horário"

                    max_linha_grade = ws.max_row
                    letra_max_col = get_column_letter(ws.max_column)

                    tabela_excel = Table(displayName=f"Grade_Fase_{fase_num}", ref=f"A1:{letra_max_col}{max_linha_grade}")
                    tabela_excel.tableStyleInfo = TableStyleInfo(
                        name="TableStyleMedium2", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False
                    )
                    ws.add_table(tabela_excel)

                    # Estilização da Legenda
                    linha_inicio_legenda = max_linha_grade + 4
                    celula_titulo = ws.cell(row=linha_inicio_legenda - 1, column=1)
                    celula_titulo.value = "Mapeamento e Legenda das Disciplinas:"
                    celula_titulo.font = Font(bold=True, size=11)

                    linhas_legenda_dados = []
                    for _, row in df_fase.iterrows():
                        prof = str(row["Professor"]).strip()
                        if not prof or prof.lower() in ["a contratar", "nan", "none", ""]:
                            prof = "?"

                        linhas_legenda_dados.append([
                            f"{row['Código da Disciplina']} ({row['Turma']})",
                            row["Nome da Disciplina"],
                            row["Local"],
                            prof
                        ])

                    df_legenda = pd.DataFrame(linhas_legenda_dados, columns=["Cód. (Turma)", "Nome da Disciplina", "Local", "Professor"])
                    df_legenda.to_excel(writer, sheet_name=f"{fase_num}ª fase", startrow=linha_inicio_legenda - 1, index=False)

                    max_linha_total = ws.max_row
                    tabela_legenda_excel = Table(displayName=f"Legenda_Fase_{fase_num}", ref=f"A{linha_inicio_legenda}:{get_column_letter(4)}{max_linha_total}")
                    tabela_legenda_excel.tableStyleInfo = TableStyleInfo(
                        name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False
                    )
                    ws.add_table(tabela_legenda_excel)

                    for col in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 13)

    print(f"[SUCESSO] Planilha com colunas 'Horário' e 'Local' gerada em: '{nome_arquivo_excel}'")