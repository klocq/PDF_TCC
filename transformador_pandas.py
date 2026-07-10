import pandas as pd
import re
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font  # Importação oficial de fontes do Openpyxl

def mapear_fase_e_tipo(codigo):
    codigo = str(codigo).strip().upper()
    
    if codigo.startswith("CIN"):
        match = re.search(r'CIN7(\d)', codigo)
        if match:
            digito_fase = match.group(1)
            if digito_fase == "9":
                return "Optativa", "Optativa"
            elif digito_fase in ["1", "2", "3", "4", "5", "6"]:
                return f"{digito_fase}ª Fase", "Obrigatória"
    
    obrigatorias_outros = {
        "CAD5103": ("1ª Fase", "Obrigatória"),
        "MTM3110": ("1ª Fase", "Obrigatória"),
        "LLV7802": ("1ª Fase", "Obrigatória"),
        "INE5111": ("4ª Fase", "Obrigatória"),
    }
    
    if codigo in obrigatorias_outros:
        return obrigatorias_outros[codigo]
        
    return "Outros / Eletiva", "Optativa"


def quebrar_horario_ufsc(string_horario):
    mapa_dias = {
        '2': 'Segunda-feira',
        '3': 'Terça-feira',
        '4': 'Quarta-feira',
        '5': 'Quinta-feira',
        '6': 'Sexta-feira',
        '7': 'Sábado'
    }
    
    blocos_matutinos = ["07:30 - 08:20", "08:20 - 09:10", "09:10 - 10:10", "10:10 - 11:00", "11:00 - 11:50"]
    blocos_vespertinos = ["13:30 - 14:20", "14:20 - 15:10", "15:10 - 16:20", "16:20 - 17:10", "17:10 - 18:00"]
    blocos_noturnos = ["18:30 - 19:20", "19:20 - 20:20", "20:20 - 21:10", "21:10 - 22:00"]
    
    horarios_explodidos = []
    sub_horarios = string_horario.split(" | ")
    
    for sub in sub_horarios:
        match = re.match(r'(\d)\.(\d{4})-(\d)', sub.strip())
        if match:
            dia_num = match.group(1)
            inicio = match.group(2)
            creditos = int(match.group(3))
            
            dia_semana = mapa_dias.get(dia_num, "Desconhecido")
            hora_formatada_inicio = f"{inicio[:2]}:{inicio[2:]}"
            
            if inicio >= "1830":
                grade_blocos = blocos_noturnos
                mapa_indices = {"18:30": 0, "19:20": 1, "20:20": 2, "21:10": 3}
            elif inicio >= "1330":
                grade_blocos = blocos_vespertinos
                mapa_indices = {"13:30": 0, "14:20": 1, "15:10": 2, "16:20": 3, "17:10": 4}
            else:
                grade_blocos = blocos_matutinos
                mapa_indices = {"07:30": 0, "08:20": 1, "09:10": 2, "10:10": 3, "11:00": 4}
            
            idx_inicio = mapa_indices.get(hora_formatada_inicio, 0)
                
            for c in range(creditos):
                if idx_inicio + c < len(grade_blocos):
                    horario_do_bloco = grade_blocos[idx_inicio + c]
                    horarios_explodidos.append({
                        "Dia": dia_semana,
                        "Horário": horario_do_bloco
                    })
                    
    return horarios_explodidos


def gerar_grade_horaria_fase(df_fase):
    todos_dias = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
    todos_horarios = [
        "07:30 - 08:20", "08:20 - 09:10", "09:10 - 10:10", "10:10 - 11:00", "11:00 - 11:50",
        "13:30 - 14:20", "14:20 - 15:10", "15:10 - 16:20", "16:20 - 17:10", "17:10 - 18:00",
        "18:30 - 19:20", "19:20 - 20:20", "20:20 - 21:10", "21:10 - 22:00"
    ]
    
    linhas_grade = []
    
    for _, linha_materia in df_fase.iterrows():
        string_horario = str(linha_materia["Horário/Local"])
        
        if string_horario and string_horario != "A definir" and string_horario != "N/A":
            blocos_tempo = quebrar_horario_ufsc(string_horario)
            for bloco in blocos_tempo:
                linhas_grade.append({
                    "Dia": bloco["Dia"],
                    "Horário": bloco["Horário"],
                    "Identificador": f"{linha_materia['Código da Disciplina']} ({linha_materia['Turma']})"
                })
                
    if not linhas_grade:
        df_vazio = pd.DataFrame("-", index=todos_horarios, columns=todos_dias)
        return df_vazio
        
    df_explodido = pd.DataFrame(linhas_grade)
    df_agrupado = df_explodido.groupby(["Horário", "Dia"])["Identificador"].apply(lambda x: " / ".join(x)).reset_index()
    
    grade_pivotada = df_agrupado.pivot(index="Horário", columns="Dia", values="Identificador")
    grade_final = grade_pivotada.reindex(index=todos_horarios, columns=todos_dias)
    grade_final = grade_final.fillna("-")
    
    return grade_final


def aplicar_transformacoes_pandas(dados_estruturados):
    print("\n[Pandas] Iniciando enriquecimento de dados com Pandas...")
    
    df = pd.DataFrame(dados_estruturados)
    
    fases_tipos = df["Código da Disciplina"].apply(mapear_fase_e_tipo)
    df["Fase"] = [item[0] for item in fases_tipos]
    df["Tipo"] = [item[1] for item in fases_tipos]
    
    colunas_ordenadas = [
        "Código da Disciplina", "Turma", "Nome da Disciplina", 
        "Fase", "Tipo", "Horas Aula", "Ofertas", "Horário/Local", "Professor"
    ]
    df = df[colunas_ordenadas]
    
    print(f"[Pandas] {len(df)} linhas processadas e enriquecidas com sucesso.")
    return df


def exportar_relatorios_finais(dados_finais, nome_arquivo_excel):
    print(f"[Pandas] Gerando arquivo Excel final com Grades e Legendas por fase...")
    
    with pd.ExcelWriter(nome_arquivo_excel, engine="openpyxl") as writer:
        dados_finais.to_excel(writer, sheet_name="Cadastro Geral", index=False)
        
        for fase_num in range(1, 7):
            fase_str = f"{fase_num}ª Fase"
            df_fase = dados_finais[dados_finais["Fase"] == fase_str]
            
            if not df_fase.empty:
                grade_semanal = gerar_grade_horaria_fase(df_fase)
                
                if not grade_semanal.empty:
                    grade_semanal.to_excel(writer, sheet_name=f"{fase_num}ª fase", index=True)
                    
                    wb = writer.book
                    ws = wb[f"{fase_num}ª fase"]
                    ws.cell(row=1, column=1).value = "Horário"
                    
                    max_linha_grade = ws.max_row
                    letra_max_col = get_column_letter(ws.max_column)
                    
                    tabela_excel = Table(displayName=f"Grade_Fase_{fase_num}", ref=f"A1:{letra_max_col}{max_linha_grade}")
                    tabela_excel.tableStyleInfo = TableStyleInfo(
                        name="TableStyleMedium2", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False
                    )
                    ws.add_table(tabela_excel)
                    
                    # --- CONFIGURAÇÃO CORRETA E LIMPA DA LEGENDA ---
                    linha_inicio_legenda = max_linha_grade + 4
                    
                    celula_titulo = ws.cell(row=linha_inicio_legenda - 1, column=1)
                    celula_titulo.value = "Mapeamento e Legenda das Disciplinas:"
                    celula_titulo.font = Font(bold=True, size=11)
                    
                    linhas_legenda_dados = []
                    for _, row in df_fase.iterrows():
                        prof = str(row["Professor"]).strip()
                        if not prof or prof.lower() in ["a contratar", "nan", "none", ""]:
                            prof = "?"
                            
                        linhas_legenda_dados.append([
                            f"{row['Código da Disciplina']} ({row['Turma']})",
                            row["Nome da Disciplina"],
                            prof
                        ])
                    
                    df_legenda = pd.DataFrame(linhas_legenda_dados, columns=["Cód. (Turma)", "Nome da Disciplina", "Professor"])
                    df_legenda.to_excel(writer, sheet_name=f"{fase_num}ª fase", startrow=linha_inicio_legenda - 1, index=False)
                    
                    max_linha_total = ws.max_row
                    tabela_legenda_excel = Table(displayName=f"Legenda_Fase_{fase_num}", ref=f"A{linha_inicio_legenda}:{get_column_letter(3)}{max_linha_total}")
                    tabela_legenda_excel.tableStyleInfo = TableStyleInfo(
                        name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False
                    )
                    ws.add_table(tabela_legenda_excel)
                    
                    for col in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
                        
    print(f"[SUCESSO] Planilha com Grades espelhadas e Legendas gerada em: '{nome_arquivo_excel}'")